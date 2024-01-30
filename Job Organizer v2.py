import os
import shutil
import zipfile
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from tqdm import tqdm

subfolers_list = ['CNC','DRAWINGS','EXCEL FILES', 'KSS','SHIPPING AND BILLING','ZIP FILES', 'temp', 'ARCHIVE']

def create_folders(job_path, subfolders_list): # function to create a subfolder within given job directory
    for subfolder in subfolders_list:
        subfolder_path = os.path.join(job_path, subfolder)
        if not os.path.exists(subfolder_path):
            os.makedirs(subfolder_path)

def extract_zip_archive_with_progress(archived_path, job_path, pbar):
    temp_folder = os.path.join(job_path, 'temp')
    if not os.path.exists(temp_folder):
        os.makedirs(temp_folder)
    total_size = os.path.getsize(archived_path)
    extracted_size = 0

    def update_progress(extracted_size):
        progress = (extracted_size/ total_size) * 100
        pbar.update(progress - pbar.n)

    with zipfile.ZipFile(archived_path, 'r') as zip_ref:
        with ThreadPoolExecutor(max_workers = 200) as exe:
            for member in zip_ref.infolist():
                    exe.submit(zip_ref.extract, member, temp_folder)
                    extracted_size += member.file_size
                    update_progress(extracted_size)

def move_files(source, destination, extensions):
    for file in os.listdir(source):
        if any(file.lower().endswith(ext) for ext in extensions):
            shutil.move(os.path.join(source, file), os.path.join(destination, file))

def index_temp(job_path, temp_path):
    for dirpath in os.walk(temp_path):
            move_files(dirpath, os.path.join(job_path, 'CNC'), ['.nc1', '.cnc', '.step', '.stp', '.dxf'])
            move_files(dirpath, os.path.join(job_path, 'DRAWINGS'), ['.pdf'])
            move_files(dirpath, os.path.join(job_path, 'ZIP FILES'), ['.zip', '.rar'])
            move_files(dirpath, os.path.join(job_path, 'SHIPPING AND BILLING'), ['master'])
            move_files(dirpath, os.path.join(job_path, 'EXCEL FILES'), ['.xlsx', '.xlsm', '.xls'])
            move_files(dirpath, os.path.join(job_path, 'KSS'), ['.kss'])

    if not os.path.exists(os.path.join(job_path, 'ARCHIVE')):
        os.makedirs(os.path.join(job_path, 'ARCHIVE'))
    shutil.copytree(temp_path, os.path.join(job_path, 'ARCHIVE'), dirs_exist_ok=True)
    shutil.rmtree(temp_path)

def organize_job(job_path, subfolders_list):
    for file in os.listdir(job_path):
        file_path = os.path.join(job_path, file)
        if os.path.isfile(file_path):
            if file.lower().endswith(('.zip','.rar')):
                print(f"Extracting '{file}'...")
                with tqdm(total=100, unit="B", unit_scale=True) as pbar:
                    extract_zip_archive_with_progress(file_path, job_path, pbar)

    index_temp(job_path, os.path.join(job_path,'temp'))

    move_files(job_path, os.path.join(job_path, 'CNC'), ['.nc1', '.cnc', '.step', '.stp', '.dxf'])
    move_files(job_path, os.path.join(job_path, 'DRAWINGS'), ['.pdf'])
    move_files(job_path, os.path.join(job_path, 'ZIP FILES'), ['.zip', '.rar'])
    move_files(job_path, os.path.join(job_path, 'SHIPPING AND BILLING'), ['master'])
    move_files(job_path, os.path.join(job_path, 'EXCEL FILES'), ['.xlsx', '.xlsm', '.xls'])
    move_files(job_path, os.path.join(job_path, 'KSS'), ['.kss'])
    
    for folder in os.listdir(job_path):
            folder_path = os.path.join(job_path, folder)
            if os.path.isdir(folder_path) and folder not in subfolders_list:
                for file in os.listdir(folder_path):
                    file_path = os.path.join(folder_path, file)
                    if os.path.isfile(file_path):
                        move_files(folder_path, os.path.join(job_path, 'DRAWINGS'), ['.pdf'])
                        move_files(folder_path, os.path.join(job_path, 'ZIP FILES'), ['.zip', '.rar'])
                        move_files(folder_path, os.path.join(job_path, 'SHIPPING AND BILLING'), ['master'])
                        move_files(folder_path, os.path.join(job_path, 'EXCEL FILES'), ['.xlsx', '.xlsm', '.xls'])
                        move_files(folder_path, os.path.join(job_path, 'CNC'), ['.nc1', '.cnc', '.step', '.stp', '.dxf'])
                        move_files(folder_path, os.path.join(job_path, 'KSS'), ['.kss'])
                shutil.move(folder_path,os.path.join(job_path,'temp'))
        
def main(subfolders_list):
    spreadsheet_path = 'data/workschedule.xlsx'
    workbook = load_workbook(spreadsheet_path)
    base_directory = 'Y:/'
    base_folder = os.path.join(base_directory, '02 JOB FILES')
    
    # Assuming your spreadsheet has columns named 'Client' and 'Job'
    client_column = 0
    job_column = 1

    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        client_name = row[client_column]
        job_name = row[job_column]
        
        if client_name is None:
                break
        
        if job_name is None:
            job_name = "Reserve"
        
        client_folder = os.path.join(base_folder, client_name.strip())
        job_folder = os.path.join(client_folder, job_name)

        if not os.path.exists(client_folder):
            response = input(f"Do you want to create a new folder for client '{client_name}'? (yes/no): ")
            if response.lower() == 'yes':
                os.makedirs(client_folder)
            else:
                existing_client = input("Enter the name of an existing client: ")
                client_folder = os.path.join(base_folder, existing_client)
                job_folder = os.path.join(client_folder, job_name)

        create_folders(job_folder, subfolders_list)

        organize_job(job_folder, subfolders_list)

if __name__ == "__main__":
    main(subfolers_list)